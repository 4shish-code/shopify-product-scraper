from __future__ import annotations

import csv
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from .columns import PROFILES
from .config import Settings
from .models import Product

BOM = "utf-8-sig"


def _b(v: bool) -> str:
    return "TRUE" if v else "FALSE"


@dataclass
class ExportStats:
    products: int = 0
    rows: int = 0
    variants: int = 0
    images: int = 0
    missing_price: int = 0
    missing_image: int = 0
    missing_sku: int = 0


def rows_shopify_import(products: list[Product], s: Settings, stats: ExportStats) -> list[dict]:
    cols = PROFILES["shopify-import"]
    blank = {c: "" for c in cols}
    rows: list[dict] = []

    for p in products:
        stats.products += 1
        if not p.images:
            stats.missing_image += 1

        opt_names = (p.option_names + ["", "", ""])[:3]
        default_only = (len(p.variants) == 1 and p.variants[0].is_default)
        seo_desc = p.plain_description[:320]

        for vi, v in enumerate(p.variants):
            stats.variants += 1
            if not v.price:
                stats.missing_price += 1
            if not v.sku:
                stats.missing_sku += 1

            r = dict(blank)
            r["URL handle"] = p.handle

            if vi == 0:
                r["Title"] = p.title
                r["Description"] = p.body_html
                r["Vendor"] = p.vendor
                r["Type"] = p.product_type
                r["Tags"] = ", ".join(p.tags)
                r["Published on online store"] = _b(s.published_status.lower() == "active")
                r["Status"] = s.published_status
                r["SEO title"] = p.title[:70]
                r["SEO description"] = seo_desc
                r["Gift card"] = "FALSE"
                if p.images:
                    img = p.images[0]
                    r["Product image URL"] = img.sized(s.image_width)
                    r["Image position"] = "1"
                    r["Image alt text"] = img.alt or p.title

            if not default_only:
                r["Option1 name"] = opt_names[0] or "Title"
                r["Option1 value"] = v.option1 or "Default Title"
                if opt_names[1]:
                    r["Option2 name"] = opt_names[1]
                    r["Option2 value"] = v.option2
                if opt_names[2]:
                    r["Option3 name"] = opt_names[2]
                    r["Option3 value"] = v.option3

            r["SKU"] = v.sku
            r["Barcode"] = v.barcode
            r["Price"] = v.price
            r["Compare-at price"] = v.compare_at_price
            r["Charge tax"] = _b(v.taxable)
            r["Inventory tracker"] = s.inventory_tracker
            r["Inventory quantity"] = str(s.default_qty_in_stock if v.available else 0)
            r["Continue selling when out of stock"] = s.inventory_policy
            r["Weight value (grams)"] = str(v.grams or 0)
            r["Weight unit for display"] = "g"
            r["Requires shipping"] = _b(v.requires_shipping)
            r["Fulfillment service"] = s.fulfillment_service
            if v.image_src:
                r["Variant image URL"] = v.image_src
            rows.append(r)

        for i, img in enumerate(p.images[1:], start=2):
            stats.images += 1
            r = dict(blank)
            r["URL handle"] = p.handle
            r["Product image URL"] = img.sized(s.image_width)
            r["Image position"] = str(i)
            r["Image alt text"] = img.alt or p.title
            rows.append(r)

        if p.images:
            stats.images += 1

    stats.rows = len(rows)
    return rows


def rows_shopify_legacy(products: list[Product], s: Settings, stats: ExportStats) -> list[dict]:
    cols = PROFILES["shopify-legacy"]
    blank = {c: "" for c in cols}
    rows: list[dict] = []

    for p in products:
        stats.products += 1
        if not p.images:
            stats.missing_image += 1
        opt_names = (p.option_names + ["", "", ""])[:3]

        for vi, v in enumerate(p.variants):
            stats.variants += 1
            if not v.price:
                stats.missing_price += 1
            if not v.sku:
                stats.missing_sku += 1

            r = dict(blank)
            r["Handle"] = p.handle
            if vi == 0:
                r["Title"] = p.title
                r["Body (HTML)"] = p.body_html
                r["Vendor"] = p.vendor
                r["Type"] = p.product_type
                r["Tags"] = ", ".join(p.tags)
                r["Published"] = _b(s.published_status.lower() == "active")
                r["Status"] = s.published_status.lower()
                r["SEO Title"] = p.title[:70]
                r["SEO Description"] = p.plain_description[:320]
                r["Gift Card"] = "FALSE"
                if p.images:
                    r["Image Src"] = p.images[0].sized(s.image_width)
                    r["Image Position"] = "1"
                    r["Image Alt Text"] = p.images[0].alt or p.title

            r["Option1 Name"] = opt_names[0] or "Title"
            r["Option1 Value"] = v.option1 or "Default Title"
            if opt_names[1]:
                r["Option2 Name"] = opt_names[1]
                r["Option2 Value"] = v.option2
            if opt_names[2]:
                r["Option3 Name"] = opt_names[2]
                r["Option3 Value"] = v.option3

            r["Variant SKU"] = v.sku
            r["Variant Barcode"] = v.barcode
            r["Variant Grams"] = str(v.grams or 0)
            r["Variant Weight Unit"] = "g"
            r["Variant Inventory Tracker"] = s.inventory_tracker
            r["Variant Inventory Qty"] = str(s.default_qty_in_stock if v.available else 0)
            r["Variant Inventory Policy"] = s.inventory_policy.lower()
            r["Variant Fulfillment Service"] = s.fulfillment_service
            r["Variant Price"] = v.price
            r["Variant Compare At Price"] = v.compare_at_price
            r["Variant Requires Shipping"] = _b(v.requires_shipping)
            r["Variant Taxable"] = _b(v.taxable)
            if v.image_src:
                r["Variant Image"] = v.image_src
            rows.append(r)

        for i, img in enumerate(p.images[1:], start=2):
            stats.images += 1
            r = dict(blank)
            r["Handle"] = p.handle
            r["Image Src"] = img.sized(s.image_width)
            r["Image Position"] = str(i)
            r["Image Alt Text"] = img.alt or p.title
            rows.append(r)
        if p.images:
            stats.images += 1

    stats.rows = len(rows)
    return rows


def rows_research(products: list[Product], s: Settings, stats: ExportStats) -> list[dict]:
    rows: list[dict] = []
    for p in products:
        stats.products += 1
        if not p.images:
            stats.missing_image += 1
        main = p.images[0].sized(s.image_width) if p.images else ""
        all_imgs = " | ".join(i.sized(s.image_width) for i in p.images)
        stats.images += len(p.images)

        for v in p.variants:
            stats.variants += 1
            if not v.price:
                stats.missing_price += 1
            if not v.sku:
                stats.missing_sku += 1
            rows.append({
                "site": p.site_key, "domain": p.site_domain, "product_id": p.id,
                "handle": p.handle, "title": p.title, "vendor": p.vendor,
                "type": p.product_type, "tags": ", ".join(p.tags),
                "variant_id": v.id, "variant_title": v.title, "sku": v.sku,
                "barcode": v.barcode, "price": v.price,
                "compare_at_price": v.compare_at_price,
                "on_sale": _b(v.on_sale), "discount_pct": v.discount_pct or "",
                "in_stock": _b(v.available), "grams": v.grams,
                "option1": v.option1, "option2": v.option2, "option3": v.option3,
                "image_count": len(p.images), "main_image": main, "all_images": all_imgs,
                "variant_image": v.image_src,
                "created_at": p.created_at, "updated_at": p.updated_at,
                "product_url": p.source_url,
                "description_text": p.plain_description[:500],
            })
    stats.rows = len(rows)
    return rows


BUILDERS = {
    "shopify-import": rows_shopify_import,
    "shopify-legacy": rows_shopify_legacy,
    "research": rows_research,
}


def build_rows(products: list[Product], profile: str, settings: Settings) -> tuple[list[dict], ExportStats]:
    if profile not in BUILDERS:
        raise ValueError(f"Unknown profile '{profile}'. Choose from: {', '.join(BUILDERS)}")
    stats = ExportStats()
    return BUILDERS[profile](products, settings, stats), stats


def validate(rows: list[dict], profile: str) -> list[str]:
    problems: list[str] = []
    if not rows:
        return ["Export is empty — no rows to write."]

    if profile.startswith("shopify"):
        hkey = "URL handle" if profile == "shopify-import" else "Handle"
        tkey = "Title"

        missing_handle = sum(1 for r in rows if not r.get(hkey))
        if missing_handle:
            problems.append(f"{missing_handle} row(s) have no {hkey} — Shopify will reject these.")

        first_seen: set[str] = set()
        dup_titles = 0
        no_title_first = 0
        for r in rows:
            h = r.get(hkey, "")
            if h not in first_seen:
                first_seen.add(h)
                if not r.get(tkey):
                    no_title_first += 1
            elif r.get(tkey):
                dup_titles += 1
        if dup_titles:
            problems.append(
                f"{dup_titles} continuation row(s) carry a Title — this creates duplicate products."
            )
        if no_title_first:
            problems.append(f"{no_title_first} product group(s) start without a Title.")

        counts = Counter(r.get(hkey, "") for r in rows)
        empty_price = sum(
            1 for r in rows
            if r.get("Price" if profile == "shopify-import" else "Variant Price") == ""
            and r.get("Product image URL" if profile == "shopify-import" else "Image Src") == ""
        )
        if empty_price:
            problems.append(f"{empty_price} row(s) have neither a price nor an image (dead rows).")
        if not counts:
            problems.append("No handles found.")

    return problems


def write_csv(rows: list[dict], path: Path, profile: str, *, delimiter: str = ",") -> Path:
    cols = PROFILES[profile]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding=BOM) as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter=delimiter,
                           extrasaction="ignore", quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    return path


def write_json(products: list[Product], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps([p.to_dict() for p in products], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def write_xlsx(rows: list[dict], path: Path, profile: str) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx export: pip install openpyxl") from exc

    cols = PROFILES[profile]
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(cols)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F6F4A")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill

    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, len(name) + 2), 44)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_image_manifest(products: list[Product], path: Path) -> tuple[Path, int]:
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with path.open("w", newline="", encoding=BOM) as f:
        w = csv.writer(f)
        w.writerow(["site", "handle", "title", "position", "alt", "image_url"])
        for p in products:
            for img in p.images:
                w.writerow([p.site_key, p.handle, p.title, img.position,
                            img.alt or p.title, img.src])
                n += 1
    return path, n


def split_by_size(rows: list[dict], max_rows: int) -> list[list[dict]]:
    if not max_rows or len(rows) <= max_rows:
        return [rows]
    hkey = "URL handle" if "URL handle" in rows[0] else "Handle"
    chunks, cur = [], []
    for r in rows:
        if len(cur) >= max_rows and r.get(hkey) != (cur[-1].get(hkey) if cur else None):
            chunks.append(cur)
            cur = []
        cur.append(r)
    if cur:
        chunks.append(cur)
    return chunks
